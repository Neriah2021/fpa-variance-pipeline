import pandas as pd
from google.cloud import bigquery
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Connect to BigQuery ─────────────────────────────────────
client = bigquery.Client(project='project-3887205e-dca0-460f-b5b')

query = """
    SELECT *
    FROM `project-3887205e-dca0-460f-b5b.Meridian_Corp.fct_pl_summary`
    ORDER BY year, month, department
"""

print("Fetching data from BigQuery...")
df = client.query(query).to_dataframe()
print(f"Loaded {len(df)} rows")

# ── Flag variances over 10% ─────────────────────────────────
df['flag'] = df['variance_pct'].apply(
    lambda x: 'OVER BUDGET' if x > 10 else ('UNDER BUDGET' if x < -10 else 'ON TRACK')
)

flagged = df[df['flag'] != 'ON TRACK']
print(f"\nFlagged items: {len(flagged)}")
print(flagged[['year', 'month', 'department', 'variance_pct', 'flag']].head(10))

# ── Write Excel report ──────────────────────────────────────
wb = openpyxl.Workbook()

# Sheet 1: Full P&L Summary
ws1 = wb.active
ws1.title = 'PL Summary'

headers = ['Year', 'Month', 'Department', 'Budget', 'Actuals', 'Variance', 'Variance %', 'Status']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(fill_type='solid', fgColor='0078D4')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row in df.iterrows():
    ws1.append([
        row['year'], row['month'], row['department'],
        round(row['total_budget'], 2), round(row['total_actuals'], 2),
        round(row['total_variance'], 2), round(row['variance_pct'], 2),
        row['budget_status']
    ])

# Sheet 2: Flagged Items
ws2 = wb.create_sheet('Flagged Items')
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(fill_type='solid', fgColor='C00000')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row in flagged.iterrows():
    ws2.append([
        row['year'], row['month'], row['department'],
        round(row['total_budget'], 2), round(row['total_actuals'], 2),
        round(row['total_variance'], 2), round(row['variance_pct'], 2),
        row['budget_status']
    ])

# Sheet 3: Department Summary
ws3 = wb.create_sheet('Dept Summary')
dept_summary = df.groupby('department').agg(
    total_budget=('total_budget', 'sum'),
    total_actuals=('total_actuals', 'sum'),
    total_variance=('total_variance', 'sum')
).reset_index()
dept_summary['variance_pct'] = (dept_summary['total_variance'] / dept_summary['total_budget'] * 100).round(2)

for col, header in enumerate(['Department', 'Total Budget', 'Total Actuals', 'Total Variance', 'Variance %'], 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(fill_type='solid', fgColor='00B294')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row in dept_summary.iterrows():
    ws3.append([
        row['department'], round(row['total_budget'], 2),
        round(row['total_actuals'], 2), round(row['total_variance'], 2),
        round(row['variance_pct'], 2)
    ])

# Auto-fit columns
for ws in [ws1, ws2, ws3]:
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

wb.save('meridian_corp_variance_report.xlsx')
print("\nExcel report saved: meridian_corp_variance_report.xlsx")